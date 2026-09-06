"""Report generators — CSV, Excel, PDF and chart exports.

Each generator takes a financial ``data`` dict (from FinancialService) and
writes bytes to a temp file, returning the path so the handler can send it via
``answer_document``.
"""

import csv
import tempfile
from datetime import datetime

from bot.services.financial import FinancialService


def _num(v) -> str:
    return f"{v:,}"


def _rows(data: dict) -> dict[str, list[tuple]]:
    """Flatten the dashboard data into tabular rows for all generators."""
    rows: dict[str, list[tuple]] = {
        "overview": [],
        "by_product": [],
        "by_config": [],
        "by_tournament": [],
        "by_category": [],
        "top_customers": [],
    }

    periods = data.get("periods", {})
    for k in ("today", "yesterday", "week", "month", "year"):
        rev = periods.get(k, 0)
        orders = periods.get(f"{k}_orders", 0)
        rows["overview"].append((f"درآمد {k}", _num(rev), f"{orders} سفارش"))
    rows["overview"].append(("درآمد کل", _num(data.get("total_revenue", 0)), ""))
    rows["overview"].append(("سفارش پرداختی", data.get("paid_orders", 0), ""))
    rows["overview"].append(("میانگین ارزش سفارش", _num(data.get("avg_order_value", 0)), ""))
    rows["overview"].append(("نرخ تبدیل", f"{round(data.get('conversion', 0) * 100, 1)}٪", ""))
    rows["overview"].append(("پرداخت در انتظار", data.get("pending_payments", 0), ""))
    rows["overview"].append(("تعداد سفارش", data.get("total_orders", 0), ""))
    for st, cnt in data.get("status_counts", {}).items():
        rows["overview"].append((f"وضعیت:{st}", cnt, ""))

    for item in data.get("by_product", []):
        rows["by_product"].append((item.get("label"), _num(item.get("units", 0)), _num(item.get("revenue", 0))))
    for item in data.get("by_config", []):
        rows["by_config"].append((item.get("label"), _num(item.get("units", 0)), _num(item.get("revenue", 0))))
    for item in data.get("by_tournament", []):
        rows["by_tournament"].append((item.get("label"), _num(item.get("orders", 0)), _num(item.get("revenue", 0))))
    for item in data.get("by_category", []):
        rows["by_category"].append((item.get("label"), _num(item.get("count", 0)), _num(item.get("revenue", 0))))
    for c in data.get("top_customers", []):
        rows["top_customers"].append(
            (f"@{c.get('username') or c.get('telegram_id')}", _num(c.get("order_count", 0)), _num(c.get("spend", 0)))
        )
    return rows


async def _data(service: FinancialService, filters: dict, include_series: bool = False) -> dict:
    d = await service.dashboard(filters)
    if include_series:
        d["series"] = await service.revenue_series(30, filters)
    return d


async def make_csv(service: FinancialService, filters: dict) -> str:
    data = await _data(service, filters)
    path = tempfile.mktemp(suffix=".csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        for table, items in _rows(data).items():
            w.writerow([f"--- {table} ---"])
            for row in items:
                w.writerow(row)
    return path


async def make_excel(service: FinancialService, filters: dict) -> str:
    from openpyxl import Workbook
    data = await _data(service, filters)
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    for table, items in _rows(data).items():
        s = wb.create_sheet(title=table[:31])
        for row in items:
            s.append(list(row))
    if data.get("series"):
        s = wb.create_sheet(title="Daily")
        s.append(["date", "revenue", "orders"])
        for d in data["series"]:
            s.append([d["date"], d["revenue"], d["orders"]])
    path = tempfile.mktemp(suffix=".xlsx")
    wb.save(path)
    return path


async def make_pdf(service: FinancialService, filters: dict) -> str:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    data = await _data(service, filters)
    path = tempfile.mktemp(suffix=".pdf")
    doc = SimpleDocTemplate(path, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("گزارش مالی", styles["Title"]),
        Spacer(1, 0.4 * cm),
        Paragraph(f"تاریخ تولید: {datetime.now():%Y-%m-%d %H:%M}", styles["Normal"]),
    ]
    for table, items in _rows(data).items():
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph(f"<b>{table}</b>", styles["Heading2"]))
        if not items:
            story.append(Paragraph("بدون داده", styles["Normal"]))
            continue
        t = Table([list(r) for r in items])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(t)
    doc.build(story)
    return path


async def make_chart(service: FinancialService, filters: dict) -> str:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    data = await _data(service, filters, include_series=True)
    series = data.get("series", [])
    path = tempfile.mktemp(suffix=".png")
    if not series:
        fig, ax = plt.subplots()
        ax.text(0.5, 0.5, "No data", ha="center")
    else:
        dates = [d["date"] for d in series]
        rev = [d["revenue"] for d in series]
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.bar(dates, rev, color="#4c9aff")
        ax.set_title("Daily Revenue")
        ax.tick_params(axis="x", rotation=90)
        fig.tight_layout()
    fig.savefig(path)
    plt.close(fig)
    return path